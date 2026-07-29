"""
tests/test_ingestion_profiling_service.py
==========================================
Focused unit/integration tests for IngestionProfilingService and
the profile_dataset() convenience function.

All tests are pure — no database, no fixtures, no async.

Covers:
  CSV / XLSX profiling, column normalisation, type inference,
  row serialization, consistency validation, exception wrapping,
  and isolation from repositories.
"""

from __future__ import annotations

import csv
import io
import math
from datetime import date, datetime
from decimal import Decimal

import openpyxl
import pytest

from app.models.data_source import FileFormat
from app.models.ingestion import InferredColumnType
from app.services.ingestion_profiling_service import (
    IngestionProfileResult,
    IngestionProfilingError,
    IngestionProfilingService,
    ProfiledColumn,
    ProfiledRow,
    profile_dataset,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _csv(headers: list[str], rows: list[list[str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _xlsx(headers: list[str], rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# Basic CSV profiling
# ===========================================================================


def test_basic_csv_profile_returns_result():
    content = _csv(["Name", "Age"], [["Alice", "30"], ["Bob", "25"]])
    result = profile_dataset(filename="test.csv", content=content)
    assert isinstance(result, IngestionProfileResult)
    assert result.original_filename == "test.csv"
    assert result.detected_file_format == FileFormat.CSV
    assert result.worksheet_name is None
    assert result.row_count == 2
    assert result.column_count == 2


def test_csv_utf8_bom_succeeds():
    base = _csv(["Score"], [["42"]])
    bom_content = b"\xef\xbb\xbf" + base
    result = profile_dataset(filename="bom.csv", content=bom_content)
    assert result.row_count == 1
    assert result.column_count == 1


def test_xlsx_first_worksheet_profiled():
    content = _xlsx(["City", "Pop"], [["Lusaka", 3360000], ["Ndola", 550000]])
    result = profile_dataset(filename="test.xlsx", content=content)
    assert result.detected_file_format == FileFormat.XLSX
    assert result.worksheet_name == "Sheet1"
    assert result.row_count == 2
    assert result.column_count == 2


# ===========================================================================
# Column normalisation
# ===========================================================================


def test_column_normalization():
    content = _csv(["Province Name", "Total (2022)"], [["Lusaka", "100"]])
    result = profile_dataset(filename="norm.csv", content=content)
    norms = [c.normalized_name for c in result.columns]
    assert norms == ["province_name", "total_2022"]


def test_duplicate_headers_deduplicated():
    content = _csv(["Value", "Value", "ID"], [["1", "2", "x"]])
    result = profile_dataset(filename="dup.csv", content=content)
    norms = [c.normalized_name for c in result.columns]
    assert len(set(norms)) == 3   # all unique after deduplication
    assert norms[0] == "value"
    assert norms[1] != norms[0]


def test_ordinal_positions_zero_based_contiguous():
    content = _csv(["A", "B", "C"], [["1", "2", "3"]])
    result = profile_dataset(filename="ord.csv", content=content)
    ordinals = [c.ordinal_position for c in result.columns]
    assert ordinals == [0, 1, 2]


def test_original_names_preserved():
    content = _csv(["Province Name", "GDP %"], [["Lusaka", "5.2"]])
    result = profile_dataset(filename="orig.csv", content=content)
    originals = [c.original_name for c in result.columns]
    assert originals == ["Province Name", "GDP %"]


# ===========================================================================
# Type inference
# ===========================================================================


def test_integer_inference():
    content = _csv(["Count"], [["1"], ["2"], ["3"]])
    result = profile_dataset(filename="int.csv", content=content)
    assert result.columns[0].inferred_type == InferredColumnType.INTEGER


def test_decimal_inference():
    content = _csv(["Value"], [["1.5"], ["2.7"]])
    result = profile_dataset(filename="dec.csv", content=content)
    assert result.columns[0].inferred_type == InferredColumnType.DECIMAL


def test_boolean_inference():
    content = _csv(["Active"], [["true"], ["false"], ["yes"]])
    result = profile_dataset(filename="bool.csv", content=content)
    assert result.columns[0].inferred_type == InferredColumnType.BOOLEAN


def test_date_inference():
    content = _csv(["Date"], [["2024-01-15"], ["2024-02-20"]])
    result = profile_dataset(filename="date.csv", content=content)
    assert result.columns[0].inferred_type == InferredColumnType.DATE


def test_datetime_inference():
    content = _csv(["TS"], [["2024-01-15T12:00:00"], ["2024-02-20T08:30:00"]])
    result = profile_dataset(filename="dt.csv", content=content)
    assert result.columns[0].inferred_type == InferredColumnType.DATETIME


def test_text_inference():
    content = _csv(["Name"], [["Alice"], ["Bob"]])
    result = profile_dataset(filename="text.csv", content=content)
    assert result.columns[0].inferred_type == InferredColumnType.TEXT


def test_leading_zero_string_infers_text():
    content = _csv(["Code"], [["00123"], ["00456"]])
    result = profile_dataset(filename="lz.csv", content=content)
    assert result.columns[0].inferred_type == InferredColumnType.TEXT


# ===========================================================================
# Column profiling — missing, nullable, unique, samples
# ===========================================================================


def test_missing_count():
    # Use two columns so one can be blank without triggering blank-row skip
    content = _csv(["X", "Y"], [["1", "a"], ["", "b"], ["3", "c"], ["", "d"]])
    result = profile_dataset(filename="miss.csv", content=content)
    col = result.columns[0]   # column X
    assert col.missing_count == 2


def test_nullable_true_when_missing():
    # Use two columns so the blank cell doesn't trigger blank-row skipping
    content = _csv(["X", "Y"], [["1", "a"], ["", "b"]])
    result = profile_dataset(filename="null.csv", content=content)
    assert result.columns[0].nullable is True


def test_nullable_false_when_no_missing():
    content = _csv(["X"], [["1"], ["2"]])
    result = profile_dataset(filename="notnull.csv", content=content)
    assert result.columns[0].nullable is False


def test_unique_count_excludes_missing():
    content = _csv(["X"], [["a"], [""], ["a"], ["b"]])
    result = profile_dataset(filename="uniq.csv", content=content)
    assert result.columns[0].unique_count == 2   # "a" and "b" (blank excluded)


def test_samples_exclude_missing():
    content = _csv(["X"], [["1"], [""], ["2"]])
    result = profile_dataset(filename="samp.csv", content=content)
    assert "" not in result.columns[0].sample_values
    assert None not in result.columns[0].sample_values


def test_sample_order_is_first_seen():
    content = _csv(["X"], [["c"], ["a"], ["b"], ["a"], ["c"]])
    result = profile_dataset(filename="order.csv", content=content)
    # c seen first, then a, then b
    samples = result.columns[0].sample_values
    assert samples[0] == "c"
    assert samples[1] == "a"
    assert samples[2] == "b"


def test_sample_limit_is_five():
    content = _csv(["X"], [[str(i)] for i in range(20)])
    result = profile_dataset(filename="lim.csv", content=content)
    assert len(result.columns[0].sample_values) <= 5


# ===========================================================================
# Row numbers
# ===========================================================================


def test_row_numbers_start_at_zero():
    content = _csv(["A"], [["1"], ["2"], ["3"]])
    result = profile_dataset(filename="rn.csv", content=content)
    numbers = [r.row_number for r in result.rows]
    assert numbers == [0, 1, 2]


def test_row_order_preserved():
    content = _csv(["A"], [["first"], ["second"], ["third"]])
    result = profile_dataset(filename="order.csv", content=content)
    values = [r.values["a"] for r in result.rows]
    assert values == ["first", "second", "third"]


# ===========================================================================
# Row serialization
# ===========================================================================


def test_missing_values_converted_to_none():
    content = _csv(["X", "Y"], [["1", ""]])
    result = profile_dataset(filename="none.csv", content=content)
    assert result.rows[0].values["y"] is None


def test_false_preserved():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Flag"])
    ws.append([False])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    result = profile_dataset(filename="false.xlsx", content=content)
    assert result.rows[0].values["flag"] is False


def test_zero_preserved():
    content = _csv(["Count"], [["0"]])
    result = profile_dataset(filename="zero.csv", content=content)
    # "0" stays as the string "0" in CSV
    assert result.rows[0].values["count"] == "0"


def test_decimal_serialized():
    """CSV decimal values are inferred as DECIMAL type and stay as strings."""
    content = _csv(["Amount"], [["12.50"]])
    result = profile_dataset(filename="dec.csv", content=content)
    val = result.rows[0].values["amount"]
    # CSV values remain as strings (decimal type is inferred, not converted)
    assert isinstance(val, str)
    assert val == "12.50"


def test_decimal_preserves_trailing_zeros():
    """CSV decimal with trailing zeros preserves the format."""
    content = _csv(["Amount"], [["100.00"]])
    result = profile_dataset(filename="dec_tz.csv", content=content)
    val = result.rows[0].values["amount"]
    assert isinstance(val, str)
    assert val == "100.00"


def test_decimal_high_precision():
    """High-precision decimal values are preserved from CSV."""
    content = _csv(["Value"], [["27.345678901234567890"]])
    result = profile_dataset(filename="dec_hp.csv", content=content)
    val = result.rows[0].values["value"]
    assert isinstance(val, str)
    assert val == "27.345678901234567890"


def test_decimal_negative():
    """Negative decimal values from CSV are preserved."""
    content = _csv(["Amount"], [["-123.45"]])
    result = profile_dataset(filename="dec_neg.csv", content=content)
    val = result.rows[0].values["amount"]
    assert isinstance(val, str)
    assert val == "-123.45"


def test_decimal_zero():
    """Decimal zero from CSV is preserved."""
    content = _csv(["Amount"], [["0.00"]])
    result = profile_dataset(filename="dec_zero.csv", content=content)
    val = result.rows[0].values["amount"]
    assert isinstance(val, str)
    assert val == "0.00"


def test_decimal_very_small():
    """Very small decimal values from CSV are preserved."""
    content = _csv(["Value"], [["0.0000000001"]])
    result = profile_dataset(filename="dec_small.csv", content=content)
    val = result.rows[0].values["value"]
    assert isinstance(val, str)
    assert val == "0.0000000001"


def test_csv_nan_string_infers_text():
    """CSV literal string "NaN" is valid TEXT (not a Decimal object).
    
    The string "NaN" from a CSV file cannot be a Decimal("NaN") object;
    it's just text. Type inference checks it as a string and rejects it
    as a decimal candidate, so it infers as TEXT.
    """
    content = _csv(["Value"], [["NaN"]])
    result = profile_dataset(filename="dec_nan.csv", content=content)
    # "NaN" fails type inference for DECIMAL (not a valid Decimal string)
    # and infers as TEXT, not DECIMAL
    assert result.columns[0].inferred_type == InferredColumnType.TEXT


def test_csv_infinity_string_infers_text():
    """CSV literal string "Infinity" is valid TEXT (not a Decimal object).
    
    The string "Infinity" from a CSV file cannot be a Decimal("Infinity") object;
    it's just text. Type inference checks it as a string and rejects it
    as a decimal candidate, so it infers as TEXT.
    """
    content = _csv(["Value"], [["Infinity"]])
    result = profile_dataset(filename="dec_inf.csv", content=content)
    # "Infinity" fails type inference for DECIMAL (not a valid Decimal string)
    # and infers as TEXT, not DECIMAL
    assert result.columns[0].inferred_type == InferredColumnType.TEXT


def test_decimal_in_samples():
    """Decimal sample values from CSV are stored as strings in column profiles."""
    content = _csv(["Amount"], [["10.50"], ["20.75"], ["30.25"]])
    result = profile_dataset(filename="dec_samples.csv", content=content)
    samples = result.columns[0].sample_values
    assert len(samples) > 0
    # All samples must be strings (CSV values are always strings)
    for sample in samples:
        assert isinstance(sample, str), f"Expected str, got {type(sample).__name__}: {sample}"


def test_direct_decimal_nan_rejected():
    """Direct Decimal NaN object is rejected during cell conversion.
    
    When XLSX parsing or other sources provide actual Decimal("NaN") objects
    (not the string "NaN"), they must be rejected by the profiling service.
    """
    from app.services.ingestion_profiling_service import _convert_cell
    
    with pytest.raises(IngestionProfilingError, match="non-finite"):
        _convert_cell("Amount", Decimal("NaN"))


def test_direct_decimal_positive_infinity_rejected():
    """Direct Decimal Infinity object is rejected during cell conversion."""
    from app.services.ingestion_profiling_service import _convert_cell
    
    with pytest.raises(IngestionProfilingError, match="non-finite"):
        _convert_cell("Amount", Decimal("Infinity"))


def test_direct_decimal_negative_infinity_rejected():
    """Direct Decimal -Infinity object is rejected during cell conversion."""
    from app.services.ingestion_profiling_service import _convert_cell
    
    with pytest.raises(IngestionProfilingError, match="non-finite"):
        _convert_cell("Amount", Decimal("-Infinity"))


def test_direct_finite_decimal_preserved():
    """Direct finite Decimal objects are preserved during cell conversion."""
    from app.services.ingestion_profiling_service import _convert_cell
    
    result = _convert_cell("Amount", Decimal("123.45"))
    # Cell conversion returns the Decimal as-is; serialization happens later
    assert isinstance(result, Decimal)
    assert result == Decimal("123.45")


def test_date_serialized_to_iso():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["EventDate"])
    ws.append([date(2024, 3, 15)])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    result = profile_dataset(filename="date.xlsx", content=content)
    val = result.rows[0].values["eventdate"]
    # openpyxl may return date as datetime(2024,3,15,0,0,0) — both ISO forms are valid
    assert isinstance(val, str)
    assert "2024-03-15" in val


def test_datetime_serialized_to_iso():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CreatedAt"])
    ws.append([datetime(2024, 3, 15, 12, 30, 0)])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    result = profile_dataset(filename="dt.xlsx", content=content)
    val = result.rows[0].values["createdat"]
    assert "2024-03-15" in val
    assert "12:30" in val


def test_nan_rejected():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Value"])
    ws.append([float("nan")])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    with pytest.raises(IngestionProfilingError):
        profile_dataset(filename="nan.xlsx", content=content)


def test_infinity_rejected():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Value"])
    ws.append([float("inf")])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    with pytest.raises(IngestionProfilingError):
        profile_dataset(filename="inf.xlsx", content=content)


def test_nested_list_rejected():
    """Parser returns str rows for CSV; test via a mock parsed dataset."""
    # Simulate by patching parse_ingestion_file to return a nested list cell
    from unittest.mock import patch
    from app.models.data_source import FileFormat
    from app.utils.ingestion_parser import ParsedDataset

    mock_parsed = ParsedDataset(
        original_column_names=["col"],
        rows=[[[1, 2, 3]]],   # nested list
        row_count=1,
        column_count=1,
        detected_file_format=FileFormat.CSV,
        worksheet_name=None,
    )
    content = _csv(["col"], [["x"]])
    with patch(
        "app.services.ingestion_profiling_service.parse_ingestion_file",
        return_value=mock_parsed,
    ):
        with pytest.raises(IngestionProfilingError):
            profile_dataset(filename="nested.csv", content=content)


def test_nested_dict_rejected():
    from unittest.mock import patch
    from app.models.data_source import FileFormat
    from app.utils.ingestion_parser import ParsedDataset

    mock_parsed = ParsedDataset(
        original_column_names=["col"],
        rows=[[{"nested": "dict"}]],
        row_count=1,
        column_count=1,
        detected_file_format=FileFormat.CSV,
        worksheet_name=None,
    )
    content = _csv(["col"], [["x"]])
    with patch(
        "app.services.ingestion_profiling_service.parse_ingestion_file",
        return_value=mock_parsed,
    ):
        with pytest.raises(IngestionProfilingError):
            profile_dataset(filename="dict.csv", content=content)


def test_bytes_rejected():
    from unittest.mock import patch
    from app.models.data_source import FileFormat
    from app.utils.ingestion_parser import ParsedDataset

    mock_parsed = ParsedDataset(
        original_column_names=["col"],
        rows=[[b"bytes"]],
        row_count=1,
        column_count=1,
        detected_file_format=FileFormat.CSV,
        worksheet_name=None,
    )
    content = _csv(["col"], [["x"]])
    with patch(
        "app.services.ingestion_profiling_service.parse_ingestion_file",
        return_value=mock_parsed,
    ):
        with pytest.raises(IngestionProfilingError):
            profile_dataset(filename="bytes.csv", content=content)


# ===========================================================================
# Consistency validation
# ===========================================================================


def test_consistency_row_count_matches():
    content = _csv(["A"], [["1"], ["2"], ["3"]])
    result = profile_dataset(filename="cons.csv", content=content)
    assert result.row_count == len(result.rows)


def test_consistency_column_count_matches():
    content = _csv(["A", "B", "C"], [["1", "2", "3"]])
    result = profile_dataset(filename="cons.csv", content=content)
    assert result.column_count == len(result.columns)


def test_consistency_every_row_has_all_columns():
    content = _csv(["X", "Y"], [["1", "2"], ["3", "4"]])
    result = profile_dataset(filename="keys.csv", content=content)
    expected = {"x", "y"}
    for row in result.rows:
        assert set(row.values.keys()) == expected


# ===========================================================================
# Parser exceptions wrapped correctly
# ===========================================================================


def test_parser_exception_wrapped_in_profiling_error():
    """An unsupported file format raises IngestionProfilingError, not the raw error."""
    with pytest.raises(IngestionProfilingError):
        profile_dataset(filename="bad.pdf", content=b"some bytes")


def test_empty_file_raises_profiling_error():
    with pytest.raises(IngestionProfilingError):
        profile_dataset(filename="empty.csv", content=b"")


def test_oversized_file_raises_profiling_error():
    huge = b"A,B\n" + b"1,2\n" * 1_000_000  # will exceed INGESTION_MAX_ROWS
    with pytest.raises(IngestionProfilingError):
        profile_dataset(filename="big.csv", content=huge)


# ===========================================================================
# Isolation — no repository or database usage
# ===========================================================================


def test_no_repository_imports_in_service():
    """The profiling service module must not import any repository classes."""
    import importlib
    import app.services.ingestion_profiling_service as svc_module

    source = importlib.util.find_spec(svc_module.__name__).origin
    with open(source, encoding="utf-8") as f:
        src = f.read()

    assert "Repository" not in src, (
        "IngestionProfilingService must not import or reference repository classes."
    )


def test_no_orm_model_creation_in_service():
    """The profiling service must not instantiate IngestionJob or DatasetColumn."""
    import importlib
    import app.services.ingestion_profiling_service as svc_module

    source = importlib.util.find_spec(svc_module.__name__).origin
    with open(source, encoding="utf-8") as f:
        src = f.read()

    assert "IngestionJob(" not in src
    assert "DatasetColumn(" not in src


def test_no_database_session_in_service():
    """The profiling service must not use AsyncSession or get_db."""
    import importlib
    import app.services.ingestion_profiling_service as svc_module

    source = importlib.util.find_spec(svc_module.__name__).origin
    with open(source, encoding="utf-8") as f:
        src = f.read()

    assert "AsyncSession" not in src
    assert "get_db" not in src
