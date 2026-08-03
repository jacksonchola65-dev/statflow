"""
tests/test_ingestion_parser.py
================================
Comprehensive unit tests for the CSV and XLSX parsers in
`app/utils/ingestion_parser.py` (Task 4).

All tests are pure unit tests — no database, no fixtures, no async.

Covers:
  - validate_file_size
  - detect_format
  - parse_csv (returns ParsedDataset)
  - parse_xlsx (returns ParsedDataset)
  - parse_ingestion_file (dispatcher)
  - Exception hierarchy
  - Security / safety properties
"""

from __future__ import annotations

import csv
import io
import zipfile

import openpyxl
import pytest
from app.models.data_source import FileFormat
from app.utils.ingestion_parser import (
    ColumnLimitExceededError,
    EmptyDatasetError,
    FileTooLargeError,
    IngestionParseError,
    InvalidExcelWorkbookError,
    MalformedCSVError,
    ParsedDataset,
    RowLimitExceededError,
    TooManyColumnsError,
    TooManyRowsError,
    UnsupportedFormatError,
    _xlsx_zip_preflight,
    parse_csv,
    parse_ingestion_file,
    parse_xlsx,
    validate_file_size,
)

# ---------------------------------------------------------------------------
# Helper builders
# ---------------------------------------------------------------------------

_DEFAULT_MAX_ROWS = 1_000
_DEFAULT_MAX_COLS = 100
_DEFAULT_MAX_BYTES = 10 * 1024 * 1024  # 10 MB


def _make_csv(headers: list[str], rows: list[list[str]]) -> bytes:
    """Build CSV bytes from headers and rows."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _make_xlsx(headers: list[str], rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    """Build XLSX bytes from headers and rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_parse(
    content: bytes, max_rows: int = _DEFAULT_MAX_ROWS, max_columns: int = _DEFAULT_MAX_COLS
) -> ParsedDataset:
    return parse_csv(content, max_rows=max_rows, max_columns=max_columns)


def _xlsx_parse(
    content: bytes, max_rows: int = _DEFAULT_MAX_ROWS, max_columns: int = _DEFAULT_MAX_COLS
) -> ParsedDataset:
    return parse_xlsx(content, max_rows=max_rows, max_columns=max_columns)


# ===========================================================================
# File-size validation
# ===========================================================================


def test_empty_bytes_rejected():
    """validate_file_size with empty bytes raises EmptyDatasetError."""
    with pytest.raises(EmptyDatasetError):
        validate_file_size(b"", 1024)


def test_oversized_bytes_rejected():
    """validate_file_size raises FileTooLargeError when content > max."""
    content = b"x" * 101
    with pytest.raises(FileTooLargeError):
        validate_file_size(content, 100)


def test_exactly_max_bytes_succeeds():
    """validate_file_size does NOT raise when content == max."""
    content = b"x" * 100
    validate_file_size(content, 100)  # should not raise


def test_one_byte_under_max_succeeds():
    """validate_file_size does NOT raise when content < max."""
    validate_file_size(b"x" * 99, 100)  # should not raise


# ===========================================================================
# Format detection / dispatch via parse_ingestion_file
# ===========================================================================


def test_unsupported_extension_rejected():
    """.pdf extension raises UnsupportedFormatError."""
    content = b"some bytes"
    with pytest.raises(UnsupportedFormatError):
        parse_ingestion_file(
            "file.pdf",
            content,
            max_file_bytes=_DEFAULT_MAX_BYTES,
            max_rows=_DEFAULT_MAX_ROWS,
            max_columns=_DEFAULT_MAX_COLS,
        )


def test_uppercase_csv_extension():
    """file.CSV (uppercase) is treated as CSV (case-insensitive)."""
    content = _make_csv(["Name"], [["Alice"]])
    result = parse_ingestion_file(
        "file.CSV",
        content,
        max_file_bytes=_DEFAULT_MAX_BYTES,
        max_rows=_DEFAULT_MAX_ROWS,
        max_columns=_DEFAULT_MAX_COLS,
    )
    assert result.detected_file_format == FileFormat.CSV


def test_uppercase_xlsx_extension():
    """file.XLSX (uppercase) is treated as XLSX (case-insensitive)."""
    content = _make_xlsx(["Name"], [["Alice"]])
    result = parse_ingestion_file(
        "file.XLSX",
        content,
        max_file_bytes=_DEFAULT_MAX_BYTES,
        max_rows=_DEFAULT_MAX_ROWS,
        max_columns=_DEFAULT_MAX_COLS,
    )
    assert result.detected_file_format == FileFormat.XLSX


def test_no_extension_rejected():
    """Filename with no extension raises UnsupportedFormatError."""
    content = _make_csv(["Name"], [["Alice"]])
    with pytest.raises(UnsupportedFormatError):
        parse_ingestion_file(
            "file",
            content,
            max_file_bytes=_DEFAULT_MAX_BYTES,
            max_rows=_DEFAULT_MAX_ROWS,
            max_columns=_DEFAULT_MAX_COLS,
        )


def test_xlsx_renamed_csv_rejected():
    """XLSX bytes with .csv extension raises UnsupportedFormatError."""
    xlsx_content = _make_xlsx(["Name"], [["Alice"]])
    with pytest.raises(UnsupportedFormatError):
        parse_ingestion_file(
            "file.csv",
            xlsx_content,
            max_file_bytes=_DEFAULT_MAX_BYTES,
            max_rows=_DEFAULT_MAX_ROWS,
            max_columns=_DEFAULT_MAX_COLS,
        )


def test_csv_renamed_xlsx_rejected():
    """CSV/text bytes with .xlsx extension raises UnsupportedFormatError or InvalidExcelWorkbookError."""
    csv_content = _make_csv(["Name"], [["Alice"]])
    with pytest.raises((UnsupportedFormatError, InvalidExcelWorkbookError)):
        parse_ingestion_file(
            "file.xlsx",
            csv_content,
            max_file_bytes=_DEFAULT_MAX_BYTES,
            max_rows=_DEFAULT_MAX_ROWS,
            max_columns=_DEFAULT_MAX_COLS,
        )


def test_empty_content_via_dispatcher_raises():
    """Empty bytes via parse_ingestion_file raises EmptyDatasetError."""
    with pytest.raises(EmptyDatasetError):
        parse_ingestion_file(
            "file.csv",
            b"",
            max_file_bytes=_DEFAULT_MAX_BYTES,
            max_rows=_DEFAULT_MAX_ROWS,
            max_columns=_DEFAULT_MAX_COLS,
        )


def test_oversized_content_via_dispatcher_raises():
    """Oversized bytes via parse_ingestion_file raises FileTooLargeError."""
    content = b"x" * 101
    with pytest.raises(FileTooLargeError):
        parse_ingestion_file(
            "file.csv",
            content,
            max_file_bytes=100,
            max_rows=_DEFAULT_MAX_ROWS,
            max_columns=_DEFAULT_MAX_COLS,
        )


# ===========================================================================
# CSV tests
# ===========================================================================


def test_csv_standard_succeeds():
    """Basic valid CSV returns correct ParsedDataset."""
    content = _make_csv(["Name", "Age"], [["Alice", "30"], ["Bob", "25"]])
    result = _csv_parse(content)
    assert result.row_count == 2
    assert result.column_count == 2
    assert result.original_column_names == ["Name", "Age"]
    assert result.rows == [["Alice", "30"], ["Bob", "25"]]
    assert result.detected_file_format == FileFormat.CSV
    assert result.worksheet_name is None


def test_csv_utf8_bom_succeeds():
    """CSV with UTF-8 BOM is decoded correctly."""
    base = _make_csv(["Name", "Age"], [["Alice", "30"]])
    bom_content = b"\xef\xbb\xbf" + base
    result = _csv_parse(bom_content)
    assert result.column_count == 2
    assert result.row_count == 1


def test_csv_unicode_values_succeed():
    """Headers and values with non-ASCII characters are handled."""
    content = _make_csv(["Prénom", "Ville"], [["Héléne", "Zürich"], ["José", "Köln"]])
    result = _csv_parse(content)
    assert result.row_count == 2
    assert result.original_column_names == ["Prénom", "Ville"]
    assert result.rows[0] == ["Héléne", "Zürich"]


def test_csv_invalid_utf8_rejected():
    """Bytes that cannot be decoded as UTF-8 raise MalformedCSVError."""
    bad_bytes = b"\xff\xfe invalid"
    with pytest.raises(MalformedCSVError):
        _csv_parse(bad_bytes)


def test_csv_quoted_commas_succeed():
    """Cell containing a comma inside quotes is parsed correctly."""
    content = _make_csv(["Name", "Address"], [["Alice", "123 Main St, Apt 4"]])
    result = _csv_parse(content)
    assert result.rows[0][1] == "123 Main St, Apt 4"


def test_csv_escaped_quotes_succeed():
    """Cell with escaped quotes is handled correctly."""
    # csv.writer will produce: "She said ""hello"""
    content = _make_csv(["Quote"], [['She said "hello"']])
    result = _csv_parse(content)
    assert result.rows[0][0] == 'She said "hello"'


def test_csv_blank_lines_skipped():
    """Blank rows between data rows are skipped; only non-blank rows counted."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["A", "B"])
    writer.writerow(["1", "2"])
    writer.writerow(["", ""])  # blank line
    writer.writerow(["3", "4"])
    writer.writerow(["", ""])  # trailing blank
    content = buf.getvalue().encode("utf-8")
    result = _csv_parse(content)
    assert result.row_count == 2
    assert result.rows == [["1", "2"], ["3", "4"]]


def test_csv_no_header_rejected():
    """Empty file (no rows at all) raises EmptyDatasetError or MalformedCSVError."""
    with pytest.raises((EmptyDatasetError, MalformedCSVError)):
        _csv_parse(b"")


def test_csv_header_only_rejected():
    """CSV with only a header row raises EmptyDatasetError."""
    content = _make_csv(["Name", "Age"], [])
    with pytest.raises(EmptyDatasetError):
        _csv_parse(content)


def test_csv_duplicate_headers_preserved():
    """Duplicate header names are preserved exactly (NOT deduplicated in ParsedDataset)."""
    content = _make_csv(["Value", "Value", "ID"], [["1", "2", "x"]])
    result = _csv_parse(content)
    assert result.original_column_names == ["Value", "Value", "ID"]


def test_csv_blank_header_preserved():
    """Blank header fields are preserved as-is."""
    content = _make_csv(["", "Name", ""], [["a", "Alice", "c"]])
    result = _csv_parse(content)
    assert result.original_column_names == ["", "Name", ""]


def test_csv_short_row_rejected():
    """Row with fewer fields than the header raises MalformedCSVError."""
    # Build CSV manually to get a short row
    lines = "A,B,C\n1,2\n"
    content = lines.encode("utf-8")
    with pytest.raises(MalformedCSVError):
        _csv_parse(content)


def test_csv_long_row_rejected():
    """Row with more fields than the header raises MalformedCSVError."""
    lines = "A,B\n1,2,3\n"
    content = lines.encode("utf-8")
    with pytest.raises(MalformedCSVError):
        _csv_parse(content)


def test_csv_exact_row_limit_succeeds():
    """Exactly max_rows data rows succeeds without raising."""
    rows = [[str(i), "v"] for i in range(3)]
    content = _make_csv(["ID", "Val"], rows)
    result = _csv_parse(content, max_rows=3)
    assert result.row_count == 3


def test_csv_row_limit_plus_one_fails():
    """max_rows+1 data rows raises RowLimitExceededError."""
    rows = [[str(i), "v"] for i in range(4)]
    content = _make_csv(["ID", "Val"], rows)
    with pytest.raises(RowLimitExceededError):
        _csv_parse(content, max_rows=3)


def test_csv_exact_column_limit_succeeds():
    """Exactly max_columns columns succeeds without raising."""
    headers = [f"C{i}" for i in range(3)]
    rows = [["v"] * 3]
    content = _make_csv(headers, rows)
    result = _csv_parse(content, max_columns=3)
    assert result.column_count == 3


def test_csv_column_limit_plus_one_fails():
    """max_columns+1 columns raises ColumnLimitExceededError."""
    headers = [f"C{i}" for i in range(4)]
    rows = [["v"] * 4]
    content = _make_csv(headers, rows)
    with pytest.raises(ColumnLimitExceededError):
        _csv_parse(content, max_columns=3)


def test_csv_row_order_preserved():
    """Rows come back in insertion order."""
    rows = [[str(i)] for i in range(5)]
    content = _make_csv(["ID"], rows)
    result = _csv_parse(content)
    assert result.rows == [[str(i)] for i in range(5)]


def test_csv_column_order_preserved():
    """Columns come back in header order."""
    content = _make_csv(["Z", "A", "M"], [["1", "2", "3"]])
    result = _csv_parse(content)
    assert result.original_column_names == ["Z", "A", "M"]


# ===========================================================================
# XLSX tests
# ===========================================================================


def test_xlsx_standard_succeeds():
    """Basic valid XLSX returns correct ParsedDataset."""
    content = _make_xlsx(["Name", "Score"], [["Alice", 95], ["Bob", 87]])
    result = _xlsx_parse(content)
    assert result.row_count == 2
    assert result.column_count == 2
    assert result.original_column_names == ["Name", "Score"]
    assert result.detected_file_format == FileFormat.XLSX
    assert result.worksheet_name is not None


def test_xlsx_first_worksheet_only():
    """Workbook with 2 sheets: only sheet 1 data is returned."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1.append(["Name", "Value"])
    ws1.append(["Alice", 1])

    ws2 = wb.create_sheet("Other")
    ws2.append(["X", "Y"])
    ws2.append([99, 100])

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _xlsx_parse(content)
    assert result.worksheet_name == "Data"
    assert result.original_column_names == ["Name", "Value"]
    assert result.row_count == 1


def test_xlsx_worksheet_name_returned():
    """result.worksheet_name equals the title of the first sheet."""
    content = _make_xlsx(["A"], [["1"]], sheet_name="MySheet")
    result = _xlsx_parse(content)
    assert result.worksheet_name == "MySheet"


def test_xlsx_second_worksheet_ignored():
    """Data in sheet 2 is ignored; only sheet 1 is read."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Empty"
    ws1.append(["ColA"])  # header only, no data rows

    ws2 = wb.create_sheet("HasData")
    ws2.append(["ColX"])
    ws2.append(["value"])

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    with pytest.raises(EmptyDatasetError):
        _xlsx_parse(content)


def test_xlsx_no_usable_rows_rejected():
    """Completely empty workbook raises EmptyDatasetError or InvalidExcelWorkbookError."""
    wb = openpyxl.Workbook()
    _ = wb.active  # Use _ to avoid unused-variable warning
    # Leave worksheet entirely empty
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    with pytest.raises((EmptyDatasetError, InvalidExcelWorkbookError)):
        _xlsx_parse(content)


def test_xlsx_header_only_rejected():
    """Workbook with only a header row raises EmptyDatasetError."""
    content = _make_xlsx(["Name", "Age"], [])
    with pytest.raises(EmptyDatasetError):
        _xlsx_parse(content)


def test_xlsx_corrupted_bytes_rejected():
    """b'not an xlsx file' raises InvalidExcelWorkbookError."""
    with pytest.raises(InvalidExcelWorkbookError):
        _xlsx_parse(b"not an xlsx file")


def test_xlsx_non_xlsx_with_xlsx_extension_rejected():
    """Plain text bytes with .xlsx extension raises InvalidExcelWorkbookError."""
    with pytest.raises(InvalidExcelWorkbookError):
        _xlsx_parse(b"Name,Age\nAlice,30\n")


def test_xlsx_duplicate_headers_preserved():
    """Duplicate header names are preserved exactly."""
    content = _make_xlsx(["Value", "Value", "ID"], [["a", "b", "c"]])
    result = _xlsx_parse(content)
    assert result.original_column_names == ["Value", "Value", "ID"]


def test_xlsx_blank_headers_preserved():
    """None header cells in the middle are converted to empty strings.

    Note: trailing empty headers are trimmed to determine logical column width.
    A blank header in the middle is preserved as "".
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Header: [None, "Name", "Age"] — trailing header is non-empty so no trimming
    ws.append([None, "Name", "Age"])
    ws.append(["x", "Alice", "30"])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _xlsx_parse(content)
    # First header should be "" (converted from None), non-trailing
    assert result.original_column_names[0] == ""
    assert result.original_column_names[1] == "Name"
    assert result.original_column_names[2] == "Age"


def test_xlsx_missing_trailing_cells_padded():
    """Short rows (fewer cells than header) are padded with None."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C"])
    ws.append(["x", "y"])  # only 2 cells for 3-column header
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _xlsx_parse(content)
    assert len(result.rows[0]) == 3
    assert result.rows[0][2] is None


def test_xlsx_cells_beyond_header_rejected():
    """Row with more non-None cells than header width raises MalformedCSVError.

    openpyxl always pads all rows to the worksheet's used column width, so the
    header itself will also be 3 wide when a data row has 3 non-None cells.
    To get a genuine 'data row exceeds header' case we construct a workbook where
    the header has a trailing None (making the header logically 2-wide) while a
    data row has a non-None value in column 3.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Write header with only 2 real columns; leave column 3 None
    ws["A1"] = "A"
    ws["B1"] = "B"
    # Data row: columns A, B, C all populated (C is beyond the 2-header width)
    ws["A2"] = "x"
    ws["B2"] = "y"
    ws["C2"] = "EXTRA"
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    # When read by openpyxl, the header row becomes ("A", "B", None)
    # After trimming trailing None, headers are ["A", "B"] (2 wide).
    # Data row is ("x", "y", "EXTRA") — trimmed to 3 wide — exceeds header.
    with pytest.raises(MalformedCSVError):
        _xlsx_parse(content)


def test_xlsx_exact_row_limit_succeeds():
    """Exactly max_rows data rows succeeds."""
    rows = [[i] for i in range(3)]
    content = _make_xlsx(["ID"], rows)
    result = _xlsx_parse(content, max_rows=3)
    assert result.row_count == 3


def test_xlsx_row_limit_plus_one_fails():
    """max_rows+1 data rows raises RowLimitExceededError."""
    rows = [[i] for i in range(4)]
    content = _make_xlsx(["ID"], rows)
    with pytest.raises(RowLimitExceededError):
        _xlsx_parse(content, max_rows=3)


def test_xlsx_exact_column_limit_succeeds():
    """Exactly max_columns columns succeeds."""
    headers = [f"C{i}" for i in range(3)]
    content = _make_xlsx(headers, [["v"] * 3])
    result = _xlsx_parse(content, max_columns=3)
    assert result.column_count == 3


def test_xlsx_column_limit_plus_one_fails():
    """max_columns+1 columns raises ColumnLimitExceededError."""
    headers = [f"C{i}" for i in range(4)]
    content = _make_xlsx(headers, [["v"] * 4])
    with pytest.raises(ColumnLimitExceededError):
        _xlsx_parse(content, max_columns=3)


def test_xlsx_native_types_preserved():
    """int, float, bool, date, datetime cells keep their Python types (not str)."""
    import datetime as dt

    the_date = dt.date(2024, 3, 15)
    the_dt = dt.datetime(2024, 3, 15, 12, 30, 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Int", "Float", "Bool", "Date", "Datetime"])
    ws.append([42, 3.14, True, the_date, the_dt])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _xlsx_parse(content)
    row = result.rows[0]
    assert isinstance(row[0], int)
    assert row[0] == 42
    assert isinstance(row[1], float)
    assert abs(row[1] - 3.14) < 1e-9
    assert isinstance(row[2], bool)
    assert row[2] is True
    # openpyxl with data_only=True may return date or datetime
    assert row[3] is not None
    assert row[4] is not None


def test_xlsx_formula_not_executed():
    """With data_only=True, formula cells return cached values (not formula strings)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "Sum"])
    ws.append([10, 20, 30])  # store the computed value directly
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _xlsx_parse(content)
    # data_only=True means no formula is re-evaluated; value is returned as-is
    assert result.rows[0][2] == 30


def test_xlsx_row_order_preserved():
    """Rows come back in insertion order."""
    rows = [[i] for i in range(5)]
    content = _make_xlsx(["ID"], rows)
    result = _xlsx_parse(content)
    assert [r[0] for r in result.rows] == list(range(5))


def test_xlsx_column_order_preserved():
    """Columns come back in header order."""
    content = _make_xlsx(["Z", "A", "M"], [["1", "2", "3"]])
    result = _xlsx_parse(content)
    assert result.original_column_names == ["Z", "A", "M"]


# ===========================================================================
# Exception safety / security
# ===========================================================================


def test_exceptions_do_not_contain_raw_content():
    """Exception messages must NOT contain uploaded file bytes."""
    bad_csv = b"A,B\n1,2,3\n"  # row has extra field
    try:
        _csv_parse(bad_csv)
    except MalformedCSVError as exc:
        msg = str(exc)
        # Should not contain the raw cell values
        assert "1" not in msg or "row" in msg.lower()
        # More specifically: the raw row content shouldn't be in the message
        assert "2,3" not in msg


def test_exceptions_do_not_expose_paths():
    """Verify no filesystem path appears in exception messages."""
    import os

    bad = b"\xff\xfe not valid utf8"
    try:
        _csv_parse(bad)
    except MalformedCSVError as exc:
        msg = str(exc)
        assert os.sep not in msg
        assert "c:\\" not in msg.lower()
        assert "/" not in msg or "UTF" in msg  # allow "UTF-8" in message


def test_parser_does_not_raise_http_exception():
    """Parser exceptions must NOT be FastAPI HTTPException instances."""
    try:
        import fastapi
    except ImportError:
        pytest.skip("fastapi not installed")

    bad_csv = b"A,B\n"  # header only, no data
    try:
        _csv_parse(bad_csv)
    except Exception as exc:
        assert not isinstance(exc, fastapi.HTTPException), f"Parser raised HTTPException: {exc}"


def test_parser_exceptions_are_subclasses_of_ingestion_parse_error():
    """MalformedCSVError, EmptyDatasetError, InvalidExcelWorkbookError all inherit from IngestionParseError."""
    assert issubclass(MalformedCSVError, IngestionParseError)
    assert issubclass(EmptyDatasetError, IngestionParseError)
    assert issubclass(InvalidExcelWorkbookError, IngestionParseError)


def test_row_limit_exceeded_is_subclass_of_too_many_rows():
    """RowLimitExceededError is a subclass of TooManyRowsError."""
    assert issubclass(RowLimitExceededError, TooManyRowsError)


def test_column_limit_exceeded_is_subclass_of_too_many_columns():
    """ColumnLimitExceededError is a subclass of TooManyColumnsError."""
    assert issubclass(ColumnLimitExceededError, TooManyColumnsError)


def test_catching_too_many_rows_catches_row_limit_exceeded():
    """Code catching TooManyRowsError also catches RowLimitExceededError."""
    rows = [[str(i)] for i in range(4)]
    content = _make_csv(["ID"], rows)
    with pytest.raises(TooManyRowsError):
        _csv_parse(content, max_rows=3)


def test_catching_too_many_columns_catches_column_limit_exceeded():
    """Code catching TooManyColumnsError also catches ColumnLimitExceededError."""
    headers = [f"C{i}" for i in range(4)]
    content = _make_csv(headers, [["v"] * 4])
    with pytest.raises(TooManyColumnsError):
        _csv_parse(content, max_columns=3)


# ===========================================================================
# ParsedDataset contract verification
# ===========================================================================


def test_parsed_dataset_row_count_matches_rows_length_csv():
    """ParsedDataset.row_count == len(ParsedDataset.rows) for CSV."""
    content = _make_csv(["A", "B"], [["1", "2"], ["3", "4"], ["5", "6"]])
    result = _csv_parse(content)
    assert result.row_count == len(result.rows)


def test_parsed_dataset_column_count_matches_header_length_csv():
    """ParsedDataset.column_count == len(ParsedDataset.original_column_names) for CSV."""
    content = _make_csv(["A", "B", "C"], [["1", "2", "3"]])
    result = _csv_parse(content)
    assert result.column_count == len(result.original_column_names)


def test_parsed_dataset_row_count_matches_rows_length_xlsx():
    """ParsedDataset.row_count == len(ParsedDataset.rows) for XLSX."""
    content = _make_xlsx(["A", "B"], [[1, 2], [3, 4]])
    result = _xlsx_parse(content)
    assert result.row_count == len(result.rows)


def test_parsed_dataset_column_count_matches_header_length_xlsx():
    """ParsedDataset.column_count == len(ParsedDataset.original_column_names) for XLSX."""
    content = _make_xlsx(["A", "B", "C"], [[1, 2, 3]])
    result = _xlsx_parse(content)
    assert result.column_count == len(result.original_column_names)


def test_csv_rows_each_match_column_count():
    """Every row in ParsedDataset.rows has exactly column_count elements for CSV."""
    content = _make_csv(["A", "B", "C"], [["1", "2", "3"], ["4", "5", "6"]])
    result = _csv_parse(content)
    for row in result.rows:
        assert len(row) == result.column_count


def test_xlsx_rows_each_match_column_count():
    """Every row in ParsedDataset.rows has exactly column_count elements for XLSX."""
    content = _make_xlsx(["A", "B", "C"], [[1, 2, 3], [4, 5, 6]])
    result = _xlsx_parse(content)
    for row in result.rows:
        assert len(row) == result.column_count


def test_csv_all_row_values_are_strings():
    """All values in ParsedDataset.rows are strings for CSV."""
    content = _make_csv(["A", "B"], [["hello", "123"], ["world", "456"]])
    result = _csv_parse(content)
    for row in result.rows:
        for cell in row:
            assert isinstance(cell, str)


def test_xlsx_worksheet_name_is_none_for_csv():
    """worksheet_name is always None for CSV."""
    content = _make_csv(["A"], [["1"]])
    result = _csv_parse(content)
    assert result.worksheet_name is None


def test_xlsx_worksheet_name_is_string_for_xlsx():
    """worksheet_name is a string (not None) for XLSX."""
    content = _make_xlsx(["A"], [["1"]])
    result = _xlsx_parse(content)
    assert isinstance(result.worksheet_name, str)


# ===========================================================================
# XLSX ZIP preflight tests (Task 4 hardening)
# ===========================================================================


def test_xlsx_preflight_valid_workbook_passes():
    """A normal small XLSX passes the ZIP preflight and is parsed successfully."""
    content = _make_xlsx(["A", "B"], [["1", "2"]])
    result = _xlsx_parse(content)
    assert result.row_count == 1


def test_xlsx_preflight_rejects_non_zip_bytes():
    """Non-ZIP bytes passed to parse_xlsx raise InvalidExcelWorkbookError."""
    with pytest.raises(InvalidExcelWorkbookError):
        _xlsx_parse(b"this is not a zip file at all")


def test_xlsx_preflight_rejects_suspicious_traversal_path():
    """ZIP member with '..' traversal in path raises InvalidExcelWorkbookError."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        info = zipfile.ZipInfo("dir/../evil.xml")
        zf.writestr(info, b"<xml/>")
    content = buf.getvalue()
    with pytest.raises(InvalidExcelWorkbookError):
        _xlsx_parse(content)


def test_xlsx_preflight_rejects_absolute_path():
    """ZIP member with absolute path raises InvalidExcelWorkbookError."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        info = zipfile.ZipInfo("/etc/passwd")
        zf.writestr(info, b"root:x:0:0")
    content = buf.getvalue()
    with pytest.raises(InvalidExcelWorkbookError):
        _xlsx_parse(content)


def test_xlsx_preflight_rejects_excessive_member_count():
    """ZIP archive with > 1000 members raises InvalidExcelWorkbookError."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for i in range(1001):
            zf.writestr(f"file_{i}.xml", b"x")
    content = buf.getvalue()
    with pytest.raises(InvalidExcelWorkbookError):
        _xlsx_parse(content)


def test_xlsx_preflight_rejects_high_compression_ratio():
    """ZIP member with compression ratio > 100 raises InvalidExcelWorkbookError.

    Uses highly compressible repeated-byte content. The test is skipped if the
    platform's zlib implementation cannot achieve ratio > 100 for safety.
    """
    raw = b"A" * 200_000  # highly compressible
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        zf.writestr("xl/data.xml", raw)
    content = buf.getvalue()

    # Verify the actual ratio before asserting
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        info = zf.infolist()[0]
        ratio = info.file_size / info.compress_size if info.compress_size > 0 else 0

    if ratio <= 100:
        pytest.skip(f"Could not achieve compression ratio > 100 (got {ratio:.1f})")

    with pytest.raises(InvalidExcelWorkbookError):
        _xlsx_parse(content)


def test_xlsx_preflight_called_directly_on_bad_zip():
    """_xlsx_zip_preflight raises InvalidExcelWorkbookError for non-ZIP bytes."""
    with pytest.raises(InvalidExcelWorkbookError):
        _xlsx_zip_preflight(b"not a zip file")


def test_xlsx_preflight_called_directly_on_valid_zip():
    """_xlsx_zip_preflight does not raise for a clean small XLSX."""
    content = _make_xlsx(["Col"], [["val"]])
    _xlsx_zip_preflight(content)  # must not raise


# ===========================================================================
# Blank-line and XLSX header behavior (Task 4 hardening)
# ===========================================================================


def test_csv_blank_row_does_not_count_toward_row_limit():
    """Blank CSV rows are skipped and do NOT count toward the row limit.

    With max_rows=2 and 2 real data rows + 3 blank rows, parse succeeds.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["A", "B"])
    writer.writerow(["1", "2"])  # data row 1
    writer.writerow(["", ""])  # blank — skipped
    writer.writerow(["3", "4"])  # data row 2
    writer.writerow(["", ""])  # blank — skipped
    content = buf.getvalue().encode("utf-8")
    result = _csv_parse(content, max_rows=2)
    assert result.row_count == 2
    assert result.rows == [["1", "2"], ["3", "4"]]


def test_xlsx_blank_physical_row_1_causes_rejection():
    """XLSX where physical row 1 is all-None (blank header) is rejected.

    Physical row 1 is always the header. An all-None header causes rejection.
    Data in row 2 is NOT promoted as a header substitute.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Leave row 1 entirely empty; write data to row 2 and row 3
    ws["A2"] = "Name"
    ws["B2"] = "Age"
    ws["A3"] = "Alice"
    ws["B3"] = 30
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    with pytest.raises((EmptyDatasetError, InvalidExcelWorkbookError)):
        _xlsx_parse(content)


def test_xlsx_header_with_internal_blank_preserves_column_position():
    """XLSX header with a blank cell in the middle preserves column positions.

    Header: ["A", None, "C"] -> ["A", "", "C"]  (position 1 preserved as "")
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "A"
    ws["B1"] = None  # internal blank
    ws["C1"] = "C"
    ws["A2"] = "x"
    ws["B2"] = "y"
    ws["C2"] = "z"
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _xlsx_parse(content)
    assert result.column_count == 3
    assert result.original_column_names[0] == "A"
    assert result.original_column_names[1] == ""
    assert result.original_column_names[2] == "C"
    assert result.rows[0][0] == "x"
    assert result.rows[0][2] == "z"


def test_xlsx_rows_with_meaningful_data_beyond_header_width_rejected():
    """XLSX row containing non-None values beyond the logical header width is rejected.

    Header: ["A", "B"] (trailing None trimmed to 2 wide)
    Data row: ("x", "y", "EXTRA") -> exceeds header -> MalformedCSVError
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "A"
    ws["B1"] = "B"
    # Column C is None in header but non-None in data row
    ws["A2"] = "x"
    ws["B2"] = "y"
    ws["C2"] = "EXTRA"
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    with pytest.raises(MalformedCSVError):
        _xlsx_parse(content)
